"""R4: watermarked Excel export, histogram + scatter viz data."""

import io

from conftest import auth, get_workspace, register_and_login, upload_csv
from openpyxl import load_workbook

CSV = ("region,amount,cost\nSouth,10,4\nSouth,20,9\nNorth,30,12\n"
       "East,40,18\nWest,50,25\n")


async def test_xlsx_export_carries_watermark(client):
    tok = await register_and_login(client)
    ws = await get_workspace(client, tok)
    ds = await upload_csv(client, tok, ws, name="sales", content=CSV)
    r = await client.get(f"/api/v1/datasets/{ds['id']}/export.xlsx",
                         headers=auth(tok))
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    sh = wb["data"]
    assert "CONFIDENTIAL" in sh.cell(1, 1).value
    assert "tenant" in sh.cell(1, 1).value          # attribution watermark
    assert sh.cell(2, 1).value == "region"          # headers on row 2
    assert sh.max_row == 2 + 5


async def test_histogram_and_scatter(client):
    tok = await register_and_login(client)
    ws = await get_workspace(client, tok)
    ds = await upload_csv(client, tok, ws, name="sales", content=CSV)
    h = (await client.get(f"/api/v1/datasets/{ds['id']}/histogram",
                          params={"column": "amount", "bins": 4},
                          headers=auth(tok))).json()
    assert len(h["bins"]) == 4
    assert sum(b["count"] for b in h["bins"]) == 5
    assert h["bins"][0]["from"] == 10 and h["max"] == 50
    r = await client.get(f"/api/v1/datasets/{ds['id']}/histogram",
                         params={"column": "region"}, headers=auth(tok))
    assert r.status_code == 422
    s = (await client.get(f"/api/v1/datasets/{ds['id']}/scatter",
                          params={"x": "amount", "y": "cost"},
                          headers=auth(tok))).json()
    assert [10.0, 4.0] in s["points"] and len(s["points"]) == 5
